import streamlit as st
import pandas as pd
import difflib
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Utility: extract all n-word phrases from a string
def get_phrases(text, min_len=2, max_len=4):
    words = re.findall(r'\w+', str(text).lower())
    phrases = set()
    for n in range(min_len, max_len + 1):
        for i in range(len(words) - n + 1):
            phrase = ' '.join(words[i:i + n])
            phrases.add(phrase)
    return phrases

# Utility: check for fuzzy matches between two sets of phrases
def fuzzy_phrase_match(set1, set2, threshold):
    for p1 in set1:
        for p2 in set2:
            ratio = difflib.SequenceMatcher(None, p1, p2).ratio()
            if ratio >= threshold:
                return True, p1, p2
    return False, None, None

# Highlight matches in workbook
def highlight_workbook(upload, min_len, max_len, threshold):
    threshold /= 100
    upload.seek(0)  # Reset pointer in case it was already read
    wb = load_workbook(upload)
    sheets = wb.sheetnames
    if len(sheets) < 2:
        return wb, []
    ws1 = wb[sheets[0]]
    ws2 = wb[sheets[1]]

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    def process_matches(ws_a, ws_b, cross=True):
        matched = []
        for row_a in ws_a.iter_rows(min_row=2, values_only=False):
            cell_a = row_a[0]
            text_a = cell_a.value
            if not text_a:
                continue
            phrases_a = get_phrases(text_a, min_len, max_len)
            for row_b in (ws_b.iter_rows(min_row=2, values_only=False) if cross else ws_a.iter_rows(min_row=2, values_only=False)):
                cell_b = row_b[0]
                text_b = cell_b.value
                if not text_b or (not cross and cell_a.coordinate == cell_b.coordinate):
                    continue
                phrases_b = get_phrases(text_b, min_len, max_len)
                matched_flag, phrase1, phrase2 = fuzzy_phrase_match(phrases_a, phrases_b, threshold)
                if matched_flag:
                    cell_a.fill = green if cross else yellow
                    cell_b.fill = green if cross else yellow
                    matched.append((text_a, text_b, phrase1, phrase2, cross))
        return matched

    summary = []
    cross_matches = process_matches(ws1, ws2, cross=True)
    # Optional: Uncomment below if you want intra-sheet matching
    # intra1_matches = process_matches(ws1, ws1, cross=False)
    # intra2_matches = process_matches(ws2, ws2, cross=False)

    summary.extend(cross_matches)  # + intra1_matches + intra2_matches)
    return wb, summary

st.set_page_config(page_title="Excel Phrase Matcher", layout="wide")
st.title("🔍 Excel Phrase Matcher with Fuzzy Logic")
st.markdown("Upload an Excel file with two sheets. The app will highlight similar rows (even if not exact) using fuzzy logic.")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

with st.sidebar:
    st.header("🔧 Matching Settings")
    min_len = st.slider("Minimum phrase length (words)", 2, 4, 3)
    max_len = st.slider("Maximum phrase length (words)", min_len, 6, 4)
    threshold = st.slider("Fuzzy Match Threshold (%)", 50, 100, 85)

if uploaded_file:
    with st.spinner("Processing file..."):
        wb, summary = highlight_workbook(uploaded_file, min_len, max_len, threshold)
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        st.success("✅ Matching completed and highlights applied!")
        st.download_button("📥 Download Highlighted Workbook", data=bio, file_name="highlighted_matches.xlsx")

        if summary:
            st.markdown("### 📊 Match Summary")
            st.dataframe(pd.DataFrame(summary, columns=["Sheet1 Row", "Sheet2 Row", "Matched Phrase 1", "Matched Phrase 2", "Cross-Sheet?"]))
        else:
            st.warning("No matches found with the current threshold.")
else:
    st.info("⬆️ Upload a file to begin.")
